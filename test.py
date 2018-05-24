import xml.etree.ElementTree as ET
import argparse
import datetime
import logging
import socket
import sys
import re
import os.path
from openpyxl import Workbook


# logging.basicConfig(format='%(asctime)s: %(levelname)s: %(message)s', filename='xml.log', filemode='w', level=logging.DEBUG)
logging.basicConfig(format='%(asctime)s: %(levelname)s: %(message)s', level=logging.DEBUG)
# tree = ET.parse('XMLFile1.xml')
tree = ET.parse('ICE.xml')
# tree = ET.parse('MSGW.xml')
xmlroot = tree.getroot()


def analytics_info(xmlroot):
    fix_acceptors = []
    network_names = []
    network_ips = []
    users = {}


    # Build dictionary
    data = {
        "core": ["Core Version", ".//Item[@name='Identity']/Item[@name='Version']"],
        "product": ["Product", ".//Item[@name='Identity']/Item[@name='Description']"],
        "hostname": ["Hostname", ".//Item[@name='Hostname']"],
        "description": ["Description", ".//Item[@name='Name']"],
        "riskport": ["Risk Port", ".//Item[@name='Risk Management']//Item[@name='Listener Port']"],
        "frapiport": ["Frapi Port", ".//Item[@name='FRAPI2']//Item[@name='Listener Port']"],
    }

    # Add fix acceptors to the dictionary
    for fix_acceptor in xmlroot.find(".//Item[@name='Client Adapters']/Item[@name='FIX']/Item[@name='Acceptors']"):
        fix_acceptors.append(fix_acceptor.attrib.get('name'))
    logging.debug(fix_acceptors)

    for acceptor in fix_acceptors:
        data["%s" % acceptor.lower()] = ["%s Port" % acceptor, ".//Item[@name='Client Adapters']/Item[@name='FIX']/Item[@name='Acceptors']//Item[@name='%s']//Item[@name='Listener Port']" % (acceptor)]

    # # Get Network IPs
    # for networks in xmlroot.find(".//Item[@name='Network']"):
    #     for network in networks.findall(".//Item[@name='IPs']"):
    #         network_ips.append(network.attrib.get('value'))

    # for networks in xmlroot.find(".//Item[@name='Network']"):
    #     # for network in networks.findall(".//Item[@name]"):
    #     #     print network.attrib.get('value')
    #     data["%s" % acceptor.lower()] = ["%s Port" % acceptor, ".//Item[@name='Client Adapters']/Item[@name='FIX']/Item[@name='Acceptors']//Item[@name='%s']//Item[@name='Listener Port']" % acceptor]

    # Add values to dictionary
    for instance in data:
        logging.debug("Getting %s", instance)
        data[instance].append(xmlroot.find(data[instance][1]).attrib.get('value'))
    for dicts in data:
        logging.debug("%s: %s", dicts, data[dicts])

    # Get Network IPs
    for network in xmlroot.findall(".//Item[@name='Network']/Item[@name]/Item[@name='Description']"):
        network_names.append(network.attrib.get('value'))
    for network in xmlroot.findall(".//Item[@name='Network']/Item[@name]/Item[@name='IPs']"):
        network_ips.append(network.attrib.get('value'))
    networks = (zip(network_names, network_ips))
    for network in networks:
        logging.debug(network)

    # Get users
    for user in xmlroot.find(".//Item[@name='Users']"):
        logging.debug("User: %s", user.attrib.get('name'))
        if user.find(".//Item[@name='Trading']//Item[@name='Orderbooks']") is not None:
            # logging.debug("Orderbook: %s", orderbook.attrib.get('name'))
            if user not in users:
                users[user.attrib.get('name')] = []
            for orderbook in user.find(".//Item[@name='Trading']//Item[@name='Orderbooks']"):
                users[user.attrib.get('name')].append(orderbook.attrib.get('name'))
                logging.debug(orderbook)




    for user in users:
        logging.debug("User %s: Orderbook %s", user, users[user])

    # If Globex run the below (currently just running as a test)
    # globex(xmlroot)

    # logging.debug(data)



    return data, networks, users


def globex(xmlroot):
    exchadapters = []
    globex = {}

    # Get exchange adapters
    for exchadapter in xmlroot.find(".//Item[@name='Exchange Adapters']"):
        exchadapters.append(exchadapter.attrib.get('name'))

    # Collect all the exchange session info into a dictionary called Globex
    for adapter in exchadapters:
        for values in xmlroot.find(".//Item[@name='Exchange Adapters']//Item[@name='%s']//Item[@name='Configuration']" % adapter):
            if adapter not in globex:
                globex[adapter] = {}
            globex[adapter][values.attrib.get('name')] = values.attrib.get('value')

    for dicts in globex:
        logging.debug("%s: %s", dicts, globex[dicts].items())

    return globex


def excel_workbook(data, exehadapter):
    data, network_ips = data
    wb = Workbook()
    sheet = wb.active
    # sheet.title = network_ips[1] + "_" + data["description"][2]
    headings = ['test', 'test2', 'test3']

    for heading in sheet.iter_rows(min_row=1):
        sheet.append(headings)



    wb.save('Handover Doc.xlsx')


def get_orderbooks(xmlroot):
    orderbooks = {}
    for grandparent in xmlroot.find(".//Item[@name='Order Management']//Item[@name='Orderbooks']"):
        orderbooks[grandparent.attrib.get('name')] = [grandparent.attrib.get('name')]

        for parent in grandparent.find(".//Item[@name='Children']"):
            orderbooks[parent.attrib.get('name')] = [parent.attrib.get('name')]
            orderbooks[grandparent.attrib.get('name')].append(parent.attrib.get('name'))

            for child in parent.find(".//Item[@name='Children']"):
                orderbooks[child.attrib.get('name')] = [child.attrib.get('name')]
                orderbooks[parent.attrib.get('name')].append(child.attrib.get('name'))
                orderbooks[grandparent.attrib.get('name')].append(child.attrib.get('name'))

                for baby in child.find(".//Item[@name='Children']"):
                    orderbooks[baby.attrib.get('name')] = [baby.attrib.get('name')]
                    orderbooks[child.attrib.get('name')].append(baby.attrib.get('name'))
                    orderbooks[parent.attrib.get('name')].append(baby.attrib.get('name'))
                    orderbooks[grandparent.attrib.get('name')].append(baby.attrib.get('name'))

    return orderbooks


def user_orderbook(xmlroot):
    orderbooks = get_orderbooks(xmlroot)
    user_accounts = {}  # dict for the user accounts and accociated parent orderbooks
    users_orderbooks = {}  # dict for users and all assigned (including child) orderbooks

    # Collects all users and accociated parent orderbooks and adds them to the user_account dict as below example
    #                                                           user_account = {amcfarlane: [amcfarlane, ne, etc..]
    for users in xmlroot.find(".//Item[@name='User Management']//Item[@name='Users']"):
        if users.find(".//Item[@name='Permissions']//Item[@name='Trading']//Item[@name='Orderbooks']"):
            user_accounts[users.attrib.get('name')] = []
            for user in users.find(".//Item[@name='Permissions']//Item[@name='Trading']//Item[@name='Orderbooks']"):
                user_accounts[users.attrib.get('name')].append(user.attrib.get('name'))

    # Diffs the orderbooks assigned to the users in the user_accounts dict against the Orderbooks dict from
    # get_orderbooks() and then adds the users with all the child orderbooks to users_orderbooks dict as below example
    #                                  user_orderbooks = {amcfarlane: [OT, amcfarlane, fpotter, ne, ne-trader1, etc..]
    for user, orderbook in user_accounts.items():  # find the user to get the orderbooks for
        users_orderbooks[user] = []
        for orderbook in user_accounts[user]:
            for ob in orderbooks[orderbook]:
                users_orderbooks[user].append(ob)

    # print example
    for user in users_orderbooks:
        logging.debug("User Orderbooks: %s %s", user, users_orderbooks[user])

    return users_orderbooks


def create_csv(data, exchadapters, orderbooks):
    logging.debug("writing csv")
    data, networks, users = data
    headings = []

    for dicts in exchadapters:
        for key in exchadapters[dicts]:
            if not key in headings:
                headings.append(key)

    f = open(data["hostname"][2]+"_"+data["description"][2]+".csv", "w")
    f.write("Hostname:,%s\n" % data['hostname'][2])
    f.write("Instance:,%s\n" % data["description"][2])
    f.write("Network IPs:\n")
    for network in networks:
        f.write("%s,%s\n" % (network[0], network[1]))
    f.write("\n")
    for lists in sorted(data.iterkeys(), reverse=True):
        if re.search(r'port|fix', lists):
            f.write(data[lists][0]+",")
    f.write("\n")
    for lists in sorted(data.iterkeys(), reverse=True):
        if re.search(r'port|fix', lists):
            f.write(data[lists][2]+",")
    f.write("\n\nExchange Adapters\n")
    for heading in headings:
        f.write("%s," % heading)
    f.write("\n")
    for adapters in exchadapters:
        for key, value in exchadapters[adapters].items():
            f.write("%s," % value)
        f.write("\n")
    f.write("\n")
    f.write("Users, Associated accounts:\n")
    for users, orderbooks in sorted(orderbooks.iteritems()):
        f.write("%s" % users)
        for orderbook in sorted(orderbooks):
            f.write(",%s" % orderbook)
        f.write("\n")







    f.write("\n")
    f.write("\n")
    f.write("\n")
    logging.debug("saving csv")
    f.close()


# def excel_workbook(data, exehadapter):
#     data, networks = data
#     wb = Workbook()
#     sheet = wb.active
#     sheet.title = data["hostname"][2]+"_"+data["description"][2]
#     # print wb.sheetnames
#
#     # sheet['A1'] = "Hostname:"
#     # sheet.append(data["hostname"][2])
#     # sheet.append("Network IPs:")
#     # for network in networks:
#     #     sheet.append(network)
#
#     sheet.append(data)
#
#
#     wb.save('Handover Doc.xlsx')


def main():
    data = analytics_info(xmlroot)
    exchadapter = globex(xmlroot)
    # excel_workbook(data, exchadapter)
    orderbooks = user_orderbook(xmlroot)
    create_csv(data, exchadapter, orderbooks)


if __name__ == '__main__':
    main()

